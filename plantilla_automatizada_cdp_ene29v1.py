import streamlit as st
import os
import fitz  # PyMuPDF
import pandas as pd
import re
from datetime import datetime
import logging
from time import sleep
import shutil
import io
from openpyxl import load_workbook

# -----------------------
# Configuración general
# -----------------------
st.set_page_config(page_title="Alcaldía Local de Usme", layout="wide")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SALIDAS_DIR = os.path.join(BASE_DIR, "salidas")
LOG_DIR = os.path.join(BASE_DIR, "logs")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
os.makedirs(SALIDAS_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(ASSETS_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, "accesos.log")
ALERTS_LOG = os.path.join(LOG_DIR, "alerts.log")

# Configurar logger con encoding utf-8 para evitar problemas futuros
logger = logging.getLogger()
logger.setLevel(logging.INFO)
if not any(isinstance(h, logging.FileHandler) and getattr(h, "baseFilename", "") == LOG_PATH for h in logger.handlers):
    handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    formatter = logging.Formatter("%(asctime)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)

# -----------------------
# Helper: rerun seguro
# -----------------------
def safe_rerun():
    """
    Intenta st.experimental_rerun(); si no está disponible, fuerza una recarga
    asignando st.query_params (API pública) y deteniendo la ejecución con st.stop().
    """
    try:
        st.experimental_rerun()
        return
    except Exception:
        pass
    try:
        current = dict(st.query_params) if getattr(st, "query_params", None) is not None else {}
    except Exception:
        current = {}
    current["_rerun"] = datetime.utcnow().isoformat() if hasattr(datetime, "utcnow") else datetime.now().isoformat()
    try:
        st.query_params = current
        st.stop()
    except Exception:
        return

# -----------------------
# Estado inicial (todas las claves usadas en session_state)
# -----------------------
if "usuario" not in st.session_state:
    st.session_state["usuario"] = None
if "df_final" not in st.session_state:
    st.session_state["df_final"] = None
if "uploaded_flag" not in st.session_state:
    st.session_state["uploaded_flag"] = False
if "processing" not in st.session_state:
    st.session_state["processing"] = False
if "auto_start" not in st.session_state:
    st.session_state["auto_start"] = False
if "auto_alerts" not in st.session_state:
    st.session_state["auto_alerts"] = False
if "roles_uploaded" not in st.session_state:
    st.session_state["roles_uploaded"] = False
if "ROLES" not in st.session_state:
    st.session_state["ROLES"] = {"admin": "admin123", "auditor": "audit456", "usuario": "user789"}
if "show_main" not in st.session_state:
    st.session_state["show_main"] = False

# -----------------------
# Utilidades específicas de extracción (basadas en tu script con fitz)
# -----------------------
def limpiar_numero(s: str) -> int:
    if not s or s in ["-", ""]:
        return 0
    s = str(s).replace("$", "").replace(" ", "").strip()
    # eliminar separadores de miles y decimales indiferenciado (heurística)
    s = s.replace(".", "").replace(",", "")
    return int(s) if s.isdigit() else 0

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto

def convertir_pep(numero: str) -> str:
    numero = str(numero).zfill(5)
    return f"PM/0005/0101/4599000{numero}"

def extraer_cdps_from_bytes(pdf_bytes: bytes, filename: str, log_lines: list) -> list:
    """
    Extrae la información esperada desde un PDF en bytes utilizando PyMuPDF (fitz).
    Devuelve una lista con un único diccionario por archivo (misma estructura que el script original).
    """
    try:
        # fitz.open puede abrir desde stream de bytes
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        log_lines.append({"Archivo": filename, "Estado": f"❌ Error abriendo PDF: {e}"})
        return []

    lineas = []
    for page in doc:
        try:
            page_text = page.get_text("text") or ""
        except Exception:
            page_text = ""
        lineas.extend(page_text.splitlines())

    valor = 0
    objeto = ""
    numero_proyecto = None
    numero_oficio = "No encontrado"
    fecha_oficio = datetime.today().strftime("%d/%m/%Y")

    for idx, line in enumerate(lineas):
        upper = line.upper() if line else ""
        # VALOR: número en la siguiente línea (heurística del script original)
        if "VALOR" in upper:
            if idx + 1 < len(lineas):
                valor_line = lineas[idx + 1]
                valor_match = re.search(r"([\d\.,]+)", valor_line)
                if valor_match:
                    valor = limpiar_numero(valor_match.group(1))

        # OBJETO: concatenar hasta encontrar VALOR
        if "OBJETO" in upper:
            objeto_lines = []
            for j in range(idx + 1, len(lineas)):
                if "VALOR" in (lineas[j].upper() if lineas[j] else ""):
                    break
                objeto_lines.append(lineas[j])
            objeto = normalizar_texto(" ".join(objeto_lines))

        # Proyecto: buscar número de 4 dígitos en línea con USME
        if "USME" in upper:
            match = re.search(r"\b(\d{4})\b", line)
            if match:
                numero_proyecto = match.group(1)

        # Solicitud No.
        if "SOLICITUD NO" in upper or "SOLICITUD N°" in upper or "SOLICITUD Nº" in upper:
            num_match = re.search(r"(\d+)", line)
            if num_match:
                numero_oficio = num_match.group(1)

        # Fecha CDP (formato YYYY/MM/DD en tu script original)
        if "CDP DE FECHA" in upper:
            fecha_match = re.search(r"(\d{4})/(\d{2})/(\d{2})", line)
            if fecha_match:
                fecha_oficio = f"{fecha_match.group(3)}/{fecha_match.group(2)}/{fecha_match.group(1)}"

    pep_convertido = convertir_pep(numero_proyecto if numero_proyecto else "0000")

    registro = {
        "Archivo": filename,
        "importe Original": valor,
        "Posición Presupuestal": "10",
        "Elemento PEP": pep_convertido,
        "Objeto": objeto,
        "Número Oficio": numero_oficio,
        "Fecha Oficio": fecha_oficio
    }

    log_lines.append({"Archivo": filename, "Estado": f"✔️ Proyecto {numero_proyecto if numero_proyecto else 'NO'} → {pep_convertido}, Valor {valor}"})
    return [registro]

# -----------------------
# Funciones de soporte (credenciales, reportes, alertas)
# -----------------------
def load_credentials_from_file(uploaded_file):
    try:
        fname = uploaded_file.name.lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        st.error(f"Error leyendo archivo de credenciales: {e}")
        return {}
    cols = {c.lower(): c for c in df.columns}
    user_col = None
    pass_col = None
    for k in cols.keys():
        if "user" in k or "usuario" in k or "login" in k:
            user_col = cols[k]
        if "pass" in k or "clave" in k or "password" in k:
            pass_col = cols[k]
    if not user_col or not pass_col:
        st.error("El archivo debe contener columnas con nombres tipo 'usuario' y 'clave' (o 'username' y 'password').")
        return {}
    roles = {}
    for _, r in df.iterrows():
        u = str(r[user_col]).strip()
        p = str(r[pass_col]).strip()
        if u:
            roles[u] = p
    return roles

def version_report():
    if not os.path.exists(LOG_PATH):
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(LOG_DIR, f"accesos_report_{ts}.log")
    try:
        shutil.copy2(LOG_PATH, dst)
        return dst
    except Exception:
        return None

def send_alert(message, level="info"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} - ALERT - {level.upper()} - {message}\n"
    try:
        with open(ALERTS_LOG, "a", encoding="utf-8") as fa:
            fa.write(line)
    except Exception:
        pass
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as fl:
            fl.write(line)
    except Exception:
        pass
    logging.info(f"ALERTA: {message}")
    return True

# -----------------------
# Estilos principales
# -----------------------
st.markdown(
    """
    <style>
    :root{
      --bg: #f6f7f9;
      --card: #ffffff;
      --muted: #6b7280;
      --text: #0f1724;
      --accent-yellow: #FFD100;
      --cta-red: #CE1126;
      --blue-accent: #0b57a4;
      --glass: rgba(15,23,36,0.04);
    }
    .stApp, .block-container, .main, .reportview-container .main .block-container {
      padding-top: 0 !important;
      margin-top: 0 !important;
      background: var(--bg) !important;
    }
    .app-wrap { padding: 8px 14px !important; margin-top: 0 !important; }
    .card { background: var(--card); border-radius: 12px; padding: 14px; box-shadow: 0 6px 18px rgba(12,18,28,0.06); }
    .stButton>button {
      border-radius: 8px; padding: 8px 14px; font-weight: 600; border: none; cursor: pointer;
      color: #fff !important; background: var(--cta-red) !important; box-shadow: 0 6px 14px rgba(206,17,38,0.12);
    }
    .stButton>button.btn-alt, .btn-alt { background: linear-gradient(90deg,var(--accent-yellow), #ffdf66) !important; color: #081226 !important; box-shadow: 0 6px 12px rgba(0,0,0,0.06) !important; }
    .stTextInput>div>div>input, .stTextInput>div>div>textarea, input[type="text"], input[type="password"] { background: #ffffff !important; color: var(--text) !important; border-radius: 6px !important; padding: 10px !important; border: 1px solid #e6e9ee !important; }
    .stFileUploader, .css-1d391kg, .css-1y4p8pa { background: linear-gradient(180deg,#ffffff,#fbfdff) !important; color: var(--text) !important; border-radius: 8px !important; padding: 10px !important; border: 1px solid #e6e9ee !important; }
    .notice-warning { background: #FFF8E1; color: #1f2933; padding: 12px 16px; border-left: 6px solid var(--accent-yellow); border-radius: 8px; margin-bottom: 8px; }
    .notice-info { background: #EAF6FF; color: #08345a; padding: 12px 16px; border-left: 6px solid var(--blue-accent); border-radius: 8px; margin-bottom: 8px; }
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------
# Layout principal
# -----------------------
st.markdown('<div class="app-wrap">', unsafe_allow_html=True)

col_left, col_main = st.columns([1, 3], gap="large")

# -----------------------
# PANEL IZQUIERDO (LOGIN + REPORTES)
# -----------------------
with col_left:
    st.markdown('<div class="card" style="position:relative;">', unsafe_allow_html=True)
    st.markdown('<div style="font-weight:700; color:var(--accent-yellow); margin-bottom:8px;">🔒 Inicio de sesión</div>', unsafe_allow_html=True)

    creds_file = st.file_uploader("Subir credenciales (Excel/CSV) — opcional", type=["xlsx", "xls", "csv"], key="creds_uploader_left")
    if creds_file is not None:
        roles_loaded = load_credentials_from_file(creds_file)
        if roles_loaded:
            st.session_state["ROLES"] = roles_loaded
            st.session_state["roles_uploaded"] = True
            st.success("Credenciales cargadas correctamente (se usarán para el inicio de sesión).")

    usuario_input = st.text_input("Usuario", key="u_input_left")
    clave_input = st.text_input("Clave", type="password", key="p_input_left")

    if st.button("Ingresar", key="btn_login_left"):
        ROLES = st.session_state.get("ROLES", {})
        if usuario_input in ROLES and ROLES[usuario_input] == clave_input:
            st.session_state["usuario"] = usuario_input
            st.session_state["show_main"] = True
            logging.info(f"Login exitoso: {usuario_input}")
        else:
            st.error("Credenciales inválidas")

    if st.session_state["usuario"] is not None:
        st.markdown(f"**Sesión activa**  \n{st.session_state['usuario']}")
        if st.button("Cerrar sesión", key="btn_logout_left"):
            st.session_state.clear()
            st.session_state["ROLES"] = {"admin": "admin123", "auditor": "audit456", "usuario": "user789"}
            st.session_state["show_main"] = False

    st.markdown("---")
    st.markdown("### 📁 Reportes & Alertas")
    if st.button("Ver Reporte de Seguridad (accesos.log)"):
        if os.path.exists(LOG_PATH):
            text = ""
            try:
                with open(LOG_PATH, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                try:
                    with open(LOG_PATH, "r", encoding="latin-1") as f:
                        text = f.read()
                except Exception:
                    with open(LOG_PATH, "rb") as f:
                        text = f.read().decode("utf-8", errors="replace")
            st.text_area("Reporte de Seguridad (accesos.log)", text, height=300)
        else:
            st.info("No hay registros de seguridad.")

    if st.button("Versionar reporte (crear copia)"):
        dst = version_report()
        if dst:
            st.success(f"Reporte versionado: {os.path.basename(dst)}")
        else:
            st.error("No se pudo versionar el reporte (¿existe accesos.log?).")

    if st.checkbox("Habilitar alertas automáticas en errores", value=st.session_state.get("auto_alerts", False), key="auto_alerts_checkbox"):
        st.session_state["auto_alerts"] = True
    else:
        st.session_state["auto_alerts"] = False

    if st.button("Enviar alerta manual (simulada)"):
        send_alert("Alerta manual enviada desde interfaz", level="warning")
        st.success("Alerta registrada y (simulada) enviada.")

    st.markdown('</div>', unsafe_allow_html=True)

# -----------------------
# PANEL PRINCIPAL (HERO + FUNCIONALIDAD)
# -----------------------
with col_main:
    st.markdown('<div class="card" style="position:relative;">', unsafe_allow_html=True)

    header_cols = st.columns([1, 5], gap="small")
    with header_cols[0]:
        img_displayed = False
        for candidate in (os.path.join(ASSETS_DIR, "descarga.jpg"),
                          os.path.join(ASSETS_DIR, "escudo.png"),
                          os.path.join(ASSETS_DIR, "escudo.jpg")):
            if os.path.exists(candidate):
                try:
                    with open(candidate, "rb") as f:
                        img_bytes = f.read()
                    st.image(img_bytes, width=140)
                    img_displayed = True
                    break
                except Exception:
                    continue

        if not img_displayed:
            st.markdown(
                """
                <div style="width:140px;height:140px;border-radius:12px;background:linear-gradient(180deg,#ffd100,#ffdf66);display:flex;align-items:center;justify-content:center;box-shadow:0 6px 18px rgba(0,0,0,0.08);">
                  <div style="width:84px;height:84px;background:#fff;border-radius:8px;display:flex;align-items:center;justify-content:center;">
                    <svg width="48" height="48" viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg" role="img" aria-label="emblema">
                      <rect width="120" height="120" rx="12" fill="#ffd100"/>
                      <path d="M36 44c8-12 20-12 28 0 8-12 20-12 28 0v8c-16-6-28 8-56 0z" fill="#000"/>
                      <circle cx="36" cy="86" r="6" fill="#d21b2b"/>
                      <circle cx="84" cy="86" r="6" fill="#d21b2b"/>
                    </svg>
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
    with header_cols[1]:
        st.markdown(
            """
            <div style="margin-left:4px;">
              <div style="font-size:28px;font-weight:800;color:#081226;">Alcaldía Local de Usme</div>
              <div style="color:#6b7280;margin-top:6px;">Gestión segura de Cargue Masivo CDP — Interfaz con identidad local</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<hr/>", unsafe_allow_html=True)

    if not st.session_state.get("show_main", False) or st.session_state.get("usuario") is None:
        st.markdown('<div class="notice-warning">Debes autenticarse en el panel izquierdo para acceder a las funciones de carga y generación de plantillas.</div>', unsafe_allow_html=True)
        st.markdown('<div class="notice-info">Después de iniciar sesión podrás subir archivos y generar la plantilla CDP.</div>', unsafe_allow_html=True)
    else:
        st.markdown("### 📊 Generador de Plantilla Cargue Masivo CDP")
        st.markdown("**Alcaldía Local de Usme**")
        st.markdown("")

        def on_upload():
            st.session_state["uploaded_flag"] = True
            st.session_state["auto_start"] = True

        st.markdown("**Sube los PDFs de contratos**  \nDrag and drop o Browse.")
        pdfs = st.file_uploader("PDFS contratos", type=["pdf"], accept_multiple_files=True, key="pdfs_main", on_change=on_upload)

        st.markdown("---")

        if st.session_state["uploaded_flag"] and not st.session_state["processing"]:
            st.markdown('<div class="notice-info">Se detectaron archivos nuevos. Pulsa Generar Plantilla o espera 2 segundos para inicio automático.</div>', unsafe_allow_html=True)
            sleep(2)

        start_now = st.button("Generar Plantilla", key="btn_generate") or st.session_state.get("auto_start", False)
        if start_now:
            st.session_state["auto_start"] = False
            if not pdfs:
                st.error("❌ Debes subir al menos un PDF.")
            else:
                st.session_state["processing"] = True
                with st.spinner("⏳ Procesando archivos..."):
                    registros = []
                    log_lines = []
                    total_pdfs = len(pdfs)
                    progress = st.progress(0)

                    for i, uploaded in enumerate(pdfs, start=1):
                        try:
                            uploaded.seek(0)
                            pdf_bytes = uploaded.read()
                            # extraer registros desde bytes
                            regs = extraer_cdps_from_bytes(pdf_bytes, uploaded.name, log_lines)
                            registros.extend(regs)
                            logging.info(f"Procesado PDF: {uploaded.name}")
                        except Exception as e:
                            msg = f"Error procesando {getattr(uploaded, 'name', str(i))}: {e}"
                            st.warning(f"⚠️ {msg}")
                            logging.error(msg)
                            log_lines.append({"Archivo": getattr(uploaded, "name", str(i)), "Estado": f"❌ {e}"})
                            if st.session_state.get("auto_alerts"):
                                send_alert(msg, level="error")
                        progress.progress(int(i / total_pdfs * 100))

                    if registros:
                        df = pd.DataFrame(registros)
                        df["CDP"] = range(1, len(df) + 1)
                        df["Num. Ext. Entidad"] = range(1, len(df) + 1)

                        fecha_actual = datetime.today().strftime("%d.%m.%Y")
                        fijos = {
                            "Posición": "1",
                            "Clase Documento": "CP",
                            "Sociedad": "1001",
                            "Moneda": "COP",
                            "Fondos": "1-100-I079",
                            "Periodo Presupuestario": "2026",
                            "Cuenta de Mayor": "7990990000",
                            "ID Solicitante": "1000131265",
                            "ID Responsable": "1000835316",
                            "Fecha Documento": fecha_actual,
                            "Fecha Contabilización": fecha_actual
                        }

                        for col, val in fijos.items():
                            df[col] = val

                        columnas_finales = [
                            "CDP", "Posición", "Fecha Documento", "Fecha Contabilización", "Clase Documento",
                            "Sociedad", "Moneda", "importe Original", "Posición Presupuestal", "Fondos",
                            "Elemento PEP", "Periodo Presupuestario", "Cuenta de Mayor", "Objeto",
                            "Número Oficio", "Fecha Oficio", "ID Solicitante", "ID Responsable",
                            "Num. Ext. Entidad", "Archivo"
                        ]

                        # Asegurar existencia de columnas antes de reordenar
                        for c in columnas_finales:
                            if c not in df.columns:
                                df[c] = ""

                        df_final = df[columnas_finales]

                        # Guardar en disk para trazabilidad y permitir descarga
                        salida_nombre = f"Plantilla_CDP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        salida_path = os.path.join(SALIDAS_DIR, salida_nombre)
                        try:
                            # Guardar DataFrame inicial
                            df_final.to_excel(salida_path, index=False, engine="openpyxl")
                        except Exception as e:
                            logging.error(f"No se pudo guardar archivo en disco: {e}")
                            if st.session_state.get("auto_alerts"):
                                send_alert(f"No se pudo guardar Excel: {e}", level="error")

                        # Añadir hoja de auditoría/log al libro existente
                        try:
                            libro = load_workbook(salida_path)
                            hoja_log = libro.create_sheet("Log_Auditoría")
                            hoja_log.append(["Archivo", "Estado"])
                            for fila in log_lines:
                                hoja_log.append([fila.get("Archivo", ""), fila.get("Estado", "")])
                            hoja_log.append([])
                            hoja_log.append(["Total PDFs procesados", total_pdfs])
                            hoja_log.append(["Registros exportados", len(df_final)])
                            libro.save(salida_path)
                        except Exception as e:
                            logging.error(f"No se pudo añadir hoja de auditoría: {e}")
                            if st.session_state.get("auto_alerts"):
                                send_alert(f"No se pudo añadir hoja de auditoría: {e}", level="error")

                        # Preparar descarga en memoria
                        towrite = io.BytesIO()
                        df_final.to_excel(towrite, index=False, engine="openpyxl")
                        towrite.seek(0)
                        st.download_button("📥 Descargar Excel", data=towrite, file_name=salida_nombre, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                        st.markdown("#### Resultado: vista previa de la plantilla generada")
                        st.dataframe(df_final, use_container_width=True)

                        # Botón volver para limpiar estado
                        if st.button("🔙 Volver"):
                            st.session_state["df_final"] = None
                            st.session_state["uploaded_flag"] = False
                            st.session_state["processing"] = False
                            for key in ("pdfs_main",):
                                if key in st.session_state:
                                    try:
                                        st.session_state[key] = None
                                    except Exception:
                                        pass
                            logging.info("Usuario pulsó Volver - retorno al estado de carga")
                            try:
                                safe_rerun()
                            except Exception:
                                try:
                                    st.experimental_rerun()
                                except Exception:
                                    pass

                        logging.info("Plantilla generada y guardada.")
                    else:
                        st.warning("⚠️ No se encontraron registros válidos en los PDFs.")
                    st.session_state["processing"] = False
                    st.session_state["uploaded_flag"] = False

    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

