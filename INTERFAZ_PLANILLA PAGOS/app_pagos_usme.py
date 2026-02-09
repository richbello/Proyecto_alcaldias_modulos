import os
import re
import io
import tempfile
import traceback
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from generador_plantilla import procesar_pagos_consolidado




# 1) PROCESO (MISMA FUNCIONALIDAD) - parametrizado por rutas
# ============================================================
def procesar_pagos_consolidado(ruta_entrada: str, ruta_destino: str) -> bool:
    # Obtener fecha actual en formato YYYYMMDD
    fecha_actual = datetime.now().strftime("%Y%m%d")
    print(f"📅 Fecha actual para columnas C y F: {fecha_actual}")

    # Leer el archivo consolidado
    print(f"Leyendo archivo: {ruta_entrada}")
    try:
        df = pd.read_excel(ruta_entrada)
        print(f"✓ Archivo leído: {len(df)} filas")

        # Mostrar las columnas que realmente tiene el archivo
        print(f"\nColumnas en el archivo:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i:2d}. {col}")

    except Exception as e:
        print(f"✗ Error al leer: {e}")
        return False

    # ===== TABLA DE EQUIVALENCIAS RETECA % -> INDICADOR =====
    # (Se conserva el comportamiento nativo de dict: claves repetidas quedan con la última asignación)
    equivalencias = {
        "0,100%": "01",
        "0,050%": "02",
        "0,200%": "03",
        "0,100%": "05",
        "0,110%": "06",
        "0,050%": "07",
        "2,000%": "08",
        "2,000%": "09",
        "0,350%": "10",
        "0,400%": "11",
        "1,000%": "12",
        "0,010%": "13",
        "0,100%": "14",
        "0,150%": "15",
        "0,250%": "16",
        "0,350%": "17",
        "0,600%": "18",
        "0,200%": "19",
        "0,250%": "20",
        "1,000%": "21",
        "1,100%": "22",
        "0,350%": "23",
        "0,600%": "24",
        "0,700%": "26",
        "0,400%": "27",
        "0,100%": "28",
        "0,200%": "29",
        "0,350%": "30",
        "0,400%": "31",
        "0,600%": "32",
        "1,104%": "33",
        "1,380%": "34",
        "0,414%": "35",
        "0,690%": "36",
        "0,700%": "37",
        "0,800%": "38",
        "0,966%": "39",
        "1,500%": "40",
        "0,250%": "41",
        "0,500%": "42",
        "0,712%": "86",
        "0,766%": "87",
        "0,866%": "88",
        "0,998%": "89",
        "1,014%": "91",
        "1,200%": "92",
        "1,214%": "R5",
        "1,400%": "94",
        "0,760%": "R3",
        "0,736%": "96",
        "1,030%": "97",
        "1,062%": "R4",
        "1,176%": "98",
        "1,254%": "99"
    }

    # Preparar la columna "Reteica %" para mapeo
    if "Reteica %" in df.columns:
        df["Reteica %"] = df["Reteica %"].astype(str)
    else:
        print("⚠ ADVERTENCIA: No se encontró la columna 'Reteica %' en el archivo consolidado")
        print("   Se buscarán columnas similares...")
        col_reteica = None
        for col in df.columns:
            if 'reteica' in str(col).lower():
                col_reteica = col
                break

        if col_reteica:
            print(f"✓ Usando columna '{col_reteica}' como fuente de porcentajes")
            df["Reteica %"] = df[col_reteica].astype(str)
        else:
            print("✗ No se encontró columna de reteica")
            df["Reteica %"] = ""

    # Mapear a indicador
    df["Indicador_Calculado"] = df["Reteica %"].map(equivalencias)

    # Mostrar algunos ejemplos del mapeo
    print(f"\n📊 Ejemplos de mapeo Reteica % -> Indicador:")
    for i in range(min(5, len(df))):
        reteica_val = df.iloc[i]["Reteica %"]
        indicador_val = df.iloc[i]["Indicador_Calculado"]
        print(f"  Pago {i+1}: '{reteica_val}' -> '{indicador_val}'")

    # Contar cuántos valores se mapearon correctamente
    mapeados = df["Indicador_Calculado"].notna().sum()
    print(f"✓ {mapeados}/{len(df)} valores mapeados a indicadores")

    # Crear archivo Excel de salida
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Encabezados exactos de la plantilla
    headers = [
        'Tipo Registro P', 'Clave Contab.', 'Codigo de la cuenta', 'Tipo Ident',
        'No Identificación', 'Indicador CME', 'Cuenta contable', 'importe',
        'Indicador de IVA', 'RP Doc Presupuestal', 'Posc Doc Pres', 'Pros Pre',
        'Programa de financiación', 'Fondo', 'Centro Gestor', 'Centro de costo',
        'Centro Beneficio', 'Orden CO', 'Elemento PEP', 'Grafo', 'Area funcional',
        'Segmento', 'Fecha Base', 'Condicion de Pago', 'Asignación', 'Texto',
        'Bloqueo Pago', 'Receptor Alternativo', 'Tipo Ident', 'No Identificación',
        'Via de Pago', 'Banco Propio', 'Id Cta', 'Ref 1', 'Ref 2', 'Referencia Pago',
        'Código Bco', 'No Cuenta', 'Tipo Cta', 'Tipo de retenciones',
        'Indicador de retención', 'Base imponible de retención', 'Importe de retención'
    ]

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = Font(bold=True)

    # Iniciar en fila 2
    fila_actual = 2

    # Procesar cada pago del consolidado
    for idx, row in df.iterrows():
        pago_num = idx + 1
        print(f"\n--- Procesando Pago {pago_num} ---")

        # 1. No Identificación (para P31)
        no_identificacion = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(word in col_lower for word in ['identific', 'nit', 'c.c', 'documento', 'cedula', 'id']):
                valor = row[col]
                if pd.notna(valor):
                    no_identificacion = str(valor).strip()
                    print(f"✓ No Identificación encontrado en '{col}': {no_identificacion}")
                    break
        if not no_identificacion:
            no_identificacion = f"ID{pago_num:04d}"
            print(f"⚠ No se encontró No Identificación, usando: {no_identificacion}")

        # 2. Valor Bruto
        valor_bruto = 0
        for col in df.columns:
            if 'valor' in str(col).lower() and 'bruto' in str(col).lower():
                valor_bruto = row[col] if pd.notna(row[col]) else 0
                print(f"✓ Valor Bruto encontrado en '{col}': {valor_bruto}")
                break

        # 3. Base imponible de retención
        base_retencion = 0
        for col in df.columns:
            col_lower = str(col).lower()
            if 'base' in col_lower and ('retencion' in col_lower or 'reteica' in col_lower):
                base_retencion = row[col] if pd.notna(row[col]) else 0
                print(f"✓ Base imponible de retención encontrada en '{col}': {base_retencion}")
                break

        # 4. Importe de retención
        importe_retencion = 0
        for col in df.columns:
            col_lower = str(col).lower()
            if 'importe' in col_lower and ('retencion' in col_lower or 'reteica' in col_lower):
                importe_retencion = row[col] if pd.notna(row[col]) else 0
                print(f"✓ Importe de retención encontrado en '{col}': {importe_retencion}")
                break
            elif 'reteica' in col_lower and 'valor' in col_lower:
                importe_retencion = row[col] if pd.notna(row[col]) else 0
                print(f"✓ Importe de retención (Reteica) encontrado en '{col}': {importe_retencion}")
                break

        # 5. RP Doc Presupuestal
        rp_doc = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(word in col_lower for word in ['rp', 'doc', 'presupuestal']):
                valor = row[col]
                if pd.notna(valor):
                    rp_doc = str(valor).strip()
                    print(f"✓ RP Doc Presupuestal encontrado en '{col}': {rp_doc}")
                    break
        if not rp_doc:
            rp_doc = f"50009973{pago_num:02d}"
            print(f"⚠ No se encontró RP Doc Presupuestal, usando: {rp_doc}")

        # 6. Asignación (número contrato)
        asignacion = ""
        for col in df.columns:
            if 'contrato' in str(col).lower():
                contrato = str(row[col]).strip() if pd.notna(row[col]) else ""
                numeros = re.findall(r'\d+', contrato)
                if numeros:
                    if len(numeros) >= 2:
                        asignacion = f"{numeros[0]}-{numeros[1]}"
                    else:
                        asignacion = numeros[0]
                print(f"✓ Contrato encontrado en '{col}': {contrato}")
                print(f"✓ Asignación generada: {asignacion}")
                break
        if not asignacion:
            asignacion = f"{pago_num:03d}-2025"
            print(f"⚠ Sin contrato, usando asignación por defecto: {asignacion}")

        # 7. Código Banco
        codigo_bco = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(word in col_lower for word in ['código', 'codigo']) and 'bco' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    codigo_bco = str(valor).strip()
                    print(f"✓ Código Banco encontrado en '{col}': {codigo_bco}")
                    break
        if not codigo_bco:
            codigo_bco = "051"
            print(f"⚠ No se encontró Código Banco, usando: {codigo_bco}")

        # 8. No Cuenta
        no_cuenta = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if 'no' in col_lower and 'cuenta' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    no_cuenta = str(valor).strip()
                    print(f"✓ No Cuenta encontrado en '{col}': {no_cuenta}")
                    break
        if not no_cuenta:
            no_cuenta = "0550488435468647"
            print(f"⚠ No se encontró No Cuenta, usando: {no_cuenta}")

        # 9. Tipo Cta
        tipo_cta = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if 'tipo' in col_lower and 'cta' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    tipo_cta = str(valor).strip()
                    print(f"✓ Tipo Cta encontrado en '{col}': {tipo_cta}")
                    break
        if not tipo_cta:
            tipo_cta = "02"
            print(f"⚠ No se encontró Tipo Cta, usando: {tipo_cta}")

        print(f"✓ Valor Bruto: {valor_bruto}")
        print(f"✓ RP Doc Presupuestal: {rp_doc}")
        print(f"✓ Base Retención (AP): {base_retencion}")
        print(f"✓ Importe Retención (AQ): {importe_retencion}")

        # ===== INDICADOR RETEICA % =====
        indicador_retencion = df.iloc[idx]["Indicador_Calculado"]
        if pd.isna(indicador_retencion) or not indicador_retencion:
            indicador_retencion = "39"
            print(f"⚠ No se encontró indicador para Reteica %, usando por defecto: {indicador_retencion}")
        else:
            print(f"✓ Indicador obtenido de Reteica %: {indicador_retencion}")

        # ===== FILA C =====
        ws.cell(row=fila_actual, column=1, value='C')
        ws.cell(row=fila_actual, column=2, value=pago_num)
        ws.cell(row=fila_actual, column=3, value=fecha_actual)
        ws.cell(row=fila_actual, column=4, value='KR')
        ws.cell(row=fila_actual, column=5, value='1001')
        ws.cell(row=fila_actual, column=6, value=fecha_actual)
        ws.cell(row=fila_actual, column=7, value='')
        ws.cell(row=fila_actual, column=8, value='COP')
        ws.cell(row=fila_actual, column=10, value=asignacion)

        # Nombre del contratista (columna K)
        nombre_contratista = ""
        for col in df.columns:
            if 'contratista' in str(col).lower():
                nombre = str(row[col]).strip() if pd.notna(row[col]) else ""
                nombre_limpio = re.sub(r'\s*(?:NIT\.|C\.C\.)\s*[\d\.,\s]+$', '', nombre).strip()
                nombre_contratista = nombre_limpio or f"CONTRATISTA {pago_num}"
                print(f"✓ Contratista encontrado: {nombre_contratista[:50]}...")
                break
        ws.cell(row=fila_actual, column=11, value=nombre_contratista)

        # ===== FILA P40 =====
        ws.cell(row=fila_actual + 1, column=1, value='P')
        ws.cell(row=fila_actual + 1, column=2, value=40)
        ws.cell(row=fila_actual + 1, column=3, value='5111809000')
        ws.cell(row=fila_actual + 1, column=4, value='')
        ws.cell(row=fila_actual + 1, column=5, value='')
        ws.cell(row=fila_actual + 1, column=8, value=valor_bruto)
        ws.cell(row=fila_actual + 1, column=9, value='WB')
        ws.cell(row=fila_actual + 1, column=10, value=rp_doc)
        ws.cell(row=fila_actual + 1, column=11, value=1)
        ws.cell(row=fila_actual + 1, column=26, value=f'10 PAGO {asignacion}')

        # ===== FILA P31 =====
        ws.cell(row=fila_actual + 2, column=1, value='P')
        ws.cell(row=fila_actual + 2, column=2, value=31)
        ws.cell(row=fila_actual + 2, column=4, value='CC')
        ws.cell(row=fila_actual + 2, column=5, value=no_identificacion)
        ws.cell(row=fila_actual + 2, column=7, value='2401010100')
        ws.cell(row=fila_actual + 2, column=8, value=valor_bruto)
        ws.cell(row=fila_actual + 2, column=24, value='0051')
        ws.cell(row=fila_actual + 2, column=25, value=asignacion)
        ws.cell(row=fila_actual + 2, column=26, value=f'10 PAGO {asignacion}')
        ws.cell(row=fila_actual + 2, column=37, value=str(codigo_bco).zfill(3))
        ws.cell(row=fila_actual + 2, column=38, value=no_cuenta)
        ws.cell(row=fila_actual + 2, column=39, value=tipo_cta)

        # ===== Indicador según Reteica % =====
        ws.cell(row=fila_actual + 2, column=40, value=indicador_retencion)  # Tipo de retenciones
        ws.cell(row=fila_actual + 2, column=41, value=indicador_retencion)  # Indicador de retención

        # Base e Importe
        ws.cell(row=fila_actual + 2, column=42, value=base_retencion)       # Base imponible
        ws.cell(row=fila_actual + 2, column=43, value=importe_retencion)    # Importe retención

        print(f"✓ Fila {fila_actual} (C): C='{fecha_actual}', E='1001', F='{fecha_actual}', J='{asignacion}'")
        print(f"✓ Fila {fila_actual+1} (P40): E='', J='{rp_doc}'")
        print(f"✓ Fila {fila_actual+2} (P31): E='{no_identificacion}'")
        print(f"✓ Fila {fila_actual+2} (P31): AN/AO='{indicador_retencion}', AP={base_retencion}, AQ={importe_retencion}")

        fila_actual += 3

    # Ajustar anchos de columnas
    anchos = {
        'A': 3, 'B': 3, 'C': 12, 'D': 3, 'E': 15, 'F': 12, 'G': 12, 'H': 10,
        'I': 3, 'J': 20, 'K': 25, 'L': 8, 'M': 25, 'N': 8, 'O': 12, 'P': 12,
        'Q': 15, 'R': 8, 'S': 12, 'T': 8, 'U': 15, 'V': 10, 'W': 10, 'X': 15,
        'Y': 12, 'Z': 30, 'AA': 12, 'AB': 20, 'AC': 3, 'AD': 15, 'AE': 10, 'AF': 12,
        'AG': 8, 'AH': 8, 'AI': 8, 'AJ': 15, 'AK': 10, 'AL': 20, 'AM': 8, 'AN': 20,
        'AO': 20, 'AP': 25, 'AQ': 20
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # Alinear texto a la izquierda
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Guardar archivo
    wb.save(ruta_destino)

    # ===== VERIFICACIÓN DE COLUMNAS CRÍTICAS =====
    print(f"\n{'='*60}")
    print("VERIFICACIÓN DE COLUMNAS CRÍTICAS")
    print('='*60)

    print(f"\n📅 Fecha usada en columnas C y F: {fecha_actual}")
    print("\nPrimeros 3 bloques (9 filas):")
    print("-" * 80)

    for fila in range(2, 11):
        valor_c = ws.cell(row=fila, column=3).value
        valor_f = ws.cell(row=fila, column=6).value
        valor_e = ws.cell(row=fila, column=5).value
        valor_j = ws.cell(row=fila, column=10).value
        valor_an = ws.cell(row=fila, column=40).value
        valor_ao = ws.cell(row=fila, column=41).value
        valor_ap = ws.cell(row=fila, column=42).value
        valor_aq = ws.cell(row=fila, column=43).value
        tipo = ws.cell(row=fila, column=1).value
        clave = ws.cell(row=fila, column=2).value

        if (fila - 2) % 3 == 0:
            tipo_fila = "C"
            c_esperado = fecha_actual
            f_esperado = fecha_actual
            e_esperado = "1001"
            j_esperado = "Asignación"
        elif (fila - 2) % 3 == 1:
            tipo_fila = "P40"
            c_esperado = "5111809000"
            f_esperado = "VACÍO"
            e_esperado = "VACÍO"
            j_esperado = "RP Doc"
        else:
            tipo_fila = "P31"
            c_esperado = "VACÍO"
            f_esperado = "VACÍO"
            e_esperado = "DATOS"
            j_esperado = "VACÍO"

        print(f"Fila {fila:2d} ({tipo_fila}): A='{tipo}', B={clave}")
        print(f"  Col C: '{valor_c}' (Esperado: {c_esperado})")
        print(f"  Col F: '{valor_f}' (Esperado: {f_esperado})")
        print(f"  Col E: '{valor_e}' (Esperado: {e_esperado})")
        print(f"  Col J (RP Doc): '{valor_j}' (Esperado: {j_esperado})")

        if tipo_fila == "P31":
            print(f"  Col AN (Tipo ret): '{valor_an}'")
            print(f"  Col AO (Ind ret): '{valor_ao}'")
            print(f"  Col AP (Base): {valor_ap}")
            print(f"  Col AQ (Importe): {valor_aq}")
        print()

    # ===== ESTADÍSTICAS =====
    print(f"\n{'='*60}")
    print("ESTADÍSTICAS DE DATOS")
    print('='*60)

    total_c = 0
    c_correctas = 0
    total_p40 = 0
    j_con_datos = 0
    total_p31 = 0
    an_con_datos = 0
    ao_con_datos = 0
    ap_con_datos = 0
    aq_con_datos = 0

    for fila in range(2, fila_actual):
        tipo = ws.cell(row=fila, column=1).value
        clave = ws.cell(row=fila, column=2).value

        if tipo == 'C':
            total_c += 1
            if ws.cell(row=fila, column=3).value == fecha_actual:
                c_correctas += 1

        if tipo == 'P' and clave == 40:
            total_p40 += 1
            if ws.cell(row=fila, column=10).value not in [None, '', ' ']:
                j_con_datos += 1

        if tipo == 'P' and clave == 31:
            total_p31 += 1
            if ws.cell(row=fila, column=40).value not in [None, '', ' ']:
                an_con_datos += 1
            if ws.cell(row=fila, column=41).value not in [None, '', ' ']:
                ao_con_datos += 1
            if ws.cell(row=fila, column=42).value not in [None, 0, '']:
                ap_con_datos += 1
            if ws.cell(row=fila, column=43).value not in [None, 0, '']:
                aq_con_datos += 1

    print(f"Total filas C: {total_c}")
    print(f"Filas C con fecha actual en columna C: {c_correctas} ({(c_correctas/total_c*100 if total_c else 0):.1f}%)")

    print(f"\nTotal filas P40: {total_p40}")
    print(f"Filas P40 con datos en J (RP Doc): {j_con_datos} ({(j_con_datos/total_p40*100 if total_p40 else 0):.1f}%)")

    print(f"\nTotal filas P31: {total_p31}")
    print(f"Filas P31 con datos en AN (Tipo ret): {an_con_datos} ({(an_con_datos/total_p31*100 if total_p31 else 0):.1f}%)")
    print(f"Filas P31 con datos en AO (Ind ret): {ao_con_datos} ({(ao_con_datos/total_p31*100 if total_p31 else 0):.1f}%)")
    print(f"Filas P31 con datos en AP (Base): {ap_con_datos} ({(ap_con_datos/total_p31*100 if total_p31 else 0):.1f}%)")
    print(f"Filas P31 con datos en AQ (Importe): {aq_con_datos} ({(aq_con_datos/total_p31*100 if total_p31 else 0):.1f}%)")

    indicadores_usados = {}
    for fila in range(2, fila_actual):
        if ws.cell(row=fila, column=1).value == 'P' and ws.cell(row=fila, column=2).value == 31:
            indicador = ws.cell(row=fila, column=40).value
            if indicador:
                indicadores_usados[indicador] = indicadores_usados.get(indicador, 0) + 1

    if indicadores_usados:
        print(f"\n📊 Indicadores de retención usados:")
        for indicador, count in sorted(indicadores_usados.items()):
            print(f"  {indicador}: {count} veces")

    print(f"\n{'='*60}")
    print(f"¡ARCHIVO GENERADO EXITOSAMENTE!")
    print(f"Ubicación: {ruta_destino}")
    print(f"Total de pagos procesados: {len(df)}")
    print(f"Total de filas generadas: {fila_actual - 1}")
    print('='*60)

    return True


# ============================================================
# 2) STREAMLIT UI (Dashboard) + Seguridad (Login + Intentos)
# ============================================================

# Colores bandera de Bogotá: Amarillo #FCDD09 y Rojo #DA121A
BOGOTA_YELLOW = "#FCDD09"
BOGOTA_RED = "#DA121A"
DARK = "#111827"
WHITE = "#FFFFFF"
BLACK = "#111111"

st.set_page_config(page_title="Alcaldía Local de Usme", layout="wide")

# --- CSS (tema bandera Bogotá + estilo dashboard)
st.markdown(f"""
<style>
/* App base */
.block-container {{
    padding-top: 0.8rem;
    padding-bottom: 1.0rem;
}}
/* Barra superior (2 franjas bandera Bogotá) */
.top-flag {{
    width: 100%;
    margin: -0.6rem 0 1rem 0;
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 4px 16px rgba(0,0,0,.10);
}}
.top-flag .y {{
    background: {BOGOTA_YELLOW};
    height: 18px;
}}
.top-flag .r {{
    background: {BOGOTA_RED};
    height: 18px;
}}

/* Sidebar */
section[data-testid="stSidebar"] > div {{
    background: {BOGOTA_YELLOW};
}}
/* Sidebar texto */
section[data-testid="stSidebar"] * {{
    color: {BLACK} !important;
}}
/* Título de sección en sidebar */
.sidebar-title {{
    background: {DARK};
    color: {WHITE} !important;
    padding: .55rem .8rem;
    border-radius: .45rem;
    font-weight: 900;
    margin-bottom: .7rem;
}}
/* Badge */
.badge {{
    display:inline-block;
    background:{BOGOTA_RED};
    color:{WHITE} !important;
    padding:.14rem .55rem;
    border-radius:999px;
    font-size:.82rem;
    font-weight:700;
}}

/* Botones (rojo Bogotá) */
.stButton > button {{
    background: {BOGOTA_RED};
    color: {WHITE};
    border: 0;
    border-radius: 10px;
    padding: .55rem .9rem;
    font-weight: 800;
}}
.stButton > button:hover {{
    filter: brightness(0.92);
}}
/* Inputs borde rojo suave */
input, textarea {{
    border-radius: 10px !important;
}}
/* Tabs: acento rojo */
button[data-baseweb="tab"] {{
    font-weight: 800 !important;
}}
button[data-baseweb="tab"][aria-selected="true"] {{
    border-bottom: 4px solid {BOGOTA_RED} !important;
}}
/* Alertas en colores consistentes */
div[data-testid="stAlert"] {{
    border-radius: 12px;
}}
</style>
""", unsafe_allow_html=True)


# --- Estado inicial
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "user" not in st.session_state:
    st.session_state.user = ""

if "failed_attempts" not in st.session_state:
    st.session_state.failed_attempts = 0
if "lock_until" not in st.session_state:
    st.session_state.lock_until = None

if "log" not in st.session_state:
    st.session_state.log = ""


# --- Login con credenciales fijas desde st.secrets
def validar_login(user: str, password: str) -> bool:
    try:
        return (user == st.secrets["APP_USER"]) and (password == st.secrets["APP_PASS"])
    except Exception:
        st.sidebar.error("⚠ Falta configurar .streamlit/secrets.toml (APP_USER / APP_PASS).")
        return False


# --- Parámetros de seguridad solicitados
MAX_INTENTOS = 5
BLOQUEO_MINUTOS = 3
MOSTRAR_INTENTOS = True


# ============================================================
# SIDEBAR (Acceso + Alertas)
# ============================================================
with st.sidebar:
    st.markdown('<div class="sidebar-title">Acceso</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([1, 2])
    with c1:
        st.markdown(
            f'<div style="width:56px;height:56px;border-radius:12px;'
            f'background:{BOGOTA_RED};display:flex;align-items:center;'
            f'justify-content:center;font-weight:900;color:{WHITE};'
            f'box-shadow:0 2px 10px rgba(0,0,0,.18);">USME</div>',
            unsafe_allow_html=True
        )
    with c2:
        st.caption("Login para acceder al panel y generar la plantilla de pagos.")

    # --- Bloqueo temporal
    if st.session_state.lock_until is not None:
        ahora = datetime.now()
        if ahora < st.session_state.lock_until:
            restante = int((st.session_state.lock_until - ahora).total_seconds())
            st.error(f"🔒 Acceso bloqueado temporalmente. Intenta de nuevo en {restante} segundos.")
            st.stop()
        else:
            st.session_state.lock_until = None
            st.session_state.failed_attempts = 0

    # --- Form login
    if not st.session_state.logged_in:
        usuario = st.text_input("Usuario", value="")
        clave = st.text_input("Clave", type="password", value="")

        if MOSTRAR_INTENTOS and st.session_state.failed_attempts > 0:
            st.warning(f"Intento {st.session_state.failed_attempts} de {MAX_INTENTOS}")

        if st.button("Ingresar", use_container_width=True):
            if validar_login(usuario.strip(), clave):
                st.session_state.logged_in = True
                st.session_state.user = usuario.strip()
                st.session_state.failed_attempts = 0
                st.success("✅ Acceso concedido")
                st.rerun()
            else:
                st.session_state.failed_attempts += 1
                restantes = MAX_INTENTOS - st.session_state.failed_attempts

                if restantes > 0:
                    st.error(f"❌ Credenciales incorrectas. Te quedan {restantes} intento(s).")
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas. Has agotado los intentos.")
                    st.session_state.lock_until = datetime.now() + timedelta(minutes=BLOQUEO_MINUTOS)
                    st.warning(f"🔒 Bloqueado por {BLOQUEO_MINUTOS} minuto(s).")
                    st.session_state.failed_attempts = 0
                    st.rerun()
    else:
        st.success(f"Sesión activa: {st.session_state.user}")
        if st.button("Cerrar sesión", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.user = ""
            st.session_state.failed_attempts = 0
            st.session_state.lock_until = None
            st.rerun()

    st.markdown("---")
    st.markdown("**Alertas automáticas**")
    st.checkbox("Ciberseguridad", value=True)
    st.checkbox("Internet", value=True)

    st.markdown("**Indicadores**")
    d1, d2 = st.columns(2)
    d1.metric("Usuarios activos", "5")
    d2.metric("Sesión", "Activa" if st.session_state.logged_in else "Inactiva")

    st.caption("Notificación local (demo)")
    st.markdown('<span class="badge">1</span> Recuperación local reset', unsafe_allow_html=True)


# ============================================================
# MAIN (Dashboard)
# ============================================================
st.markdown(
    '<div class="top-flag"><div class="y"></div><div class="r"></div></div>',
    unsafe_allow_html=True
)

st.title("Alcaldía Local de Usme")
st.caption("Panel de control • Seguridad / Audio / Bodega")

tab1, tab2, tab3 = st.tabs(["🛡️ Seguridad", "🔊 Audio", "📦 Bodega"])

# ---------------- TAB 1: Seguridad ----------------
with tab1:
    st.subheader("👋 Bienvenido")
    st.write("Inicia sesión desde el panel izquierdo.")
    st.markdown("""
- **Admin - Persona - Auditar - Administración**
- **Usuarios - Personas - Auditoría**
- **Auditor - Solo Auditoría**
""")

    st.divider()
    st.subheader("🧾 Generador de plantilla de pagos (Consolidado ➜ Plantilla)")

    # ---- BLOQUEO POR LOGIN
    if not st.session_state.logged_in:
        st.warning("🔒 Debes iniciar sesión para acceder al generador de plantilla.")
    else:
        st.success("✅ Acceso habilitado. Puedes generar la plantilla.")

        st.info("📌 Sube el consolidado (.xlsx). Al finalizar podrás descargar el Excel generado.")

        up = st.file_uploader("📤 Subir consolidado", type=["xlsx", "xls"])

        nombre_salida = st.text_input(
            "Nombre del Excel de salida",
            value=f"V1_PLANTILLA_PAGOS_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

        colA, colB = st.columns([1, 1])
        ejecutar = colA.button("▶ Generar plantilla", use_container_width=True, disabled=(up is None))
        limpiar = colB.button("🧹 Limpiar log", use_container_width=True)

        if limpiar:
            st.session_state.log = ""

        log_area = st.empty()
        log_area.text_area("Log / consola", st.session_state.log, height=320)

        if ejecutar and up is not None:
            with st.spinner("Procesando..."):
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        ruta_in = os.path.join(tmpdir, up.name)
                        with open(ruta_in, "wb") as f:
                            f.write(up.getbuffer())

                        ruta_out = os.path.join(tmpdir, nombre_salida)

                        # Capturar prints del proceso
                        buffer = io.StringIO()
                        old_stdout = os.sys.stdout
                        os.sys.stdout = buffer

                        ok = False
                        try:
                            print("=" * 60)
                            print("GENERADOR DE PLANTILLA DE PAGOS - CON FECHA ACTUAL Y RETECA %")
                            print("=" * 60)
                            ok = procesar_pagos_consolidado(ruta_in, ruta_out)
                        finally:
                            os.sys.stdout = old_stdout

                        st.session_state.log += buffer.getvalue()
                        log_area.text_area("Log / consola", st.session_state.log, height=320)

                        if ok and os.path.exists(ruta_out):
                            st.success("✅ Plantilla generada correctamente.")
                            with open(ruta_out, "rb") as f:
                                st.download_button(
                                    "⬇️ Descargar Excel generado",
                                    data=f,
                                    file_name=nombre_salida,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                        else:
                            st.error("❌ No se generó el archivo. Revisa el log.")

                except Exception as e:
                    st.session_state.log += "\n❌ Error inesperado:\n" + str(e) + "\n"
                    st.session_state.log += traceback.format_exc() + "\n"
                    log_area.text_area("Log / consola", st.session_state.log, height=320)
                    st.error("❌ Error inesperado. Mira el log.")

# ---------------- TAB 2: Audio ----------------
with tab2:
    st.subheader("🔊 Audio")
    st.write("Módulo placeholder para mantener la interfaz igual.")

# ---------------- TAB 3: Bodega ----------------
with tab3:
    st.subheader("📦 Bodega")
    st.write("Módulo placeholder para mantener la interfaz igual.")